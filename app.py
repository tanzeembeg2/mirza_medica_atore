from __future__ import annotations

import io
import json
import os
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


BASE_DIR = Path(__file__).resolve().parent
DATA_PATH = Path(os.getenv("DATA_PATH", BASE_DIR / "pharmacy_data.json"))

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "mirza-medical-store-secret")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def seed_data() -> dict:
    today = date.today()
    return {
        "next_ids": {"medicine": 5, "customer": 2, "sale": 1, "sale_item": 1, "user": 2},
        "pharmacy": {
            "name": "Mirza Medical Store",
            "address": "Gopamau Hardoi",
            "phone": "",
            "owner_name": "Admin",
        },
        "users": [
            {
                "id": 1,
                "username": "admin",
                "password": "1234",
                "role": "admin",
                "display_name": "Admin",
                "phone": "",
                "active": True,
                "created_at": now_iso(),
            }
        ],
        "medicines": [
            {"id": 1, "code": "MED-001", "name": "Paracetamol 650", "category": "Tablet", "batch_no": "PCM2401", "expiry_date": (today + timedelta(days=420)).isoformat(), "supplier": "Sun Pharma", "unit": "Strip", "quantity": 120, "min_stock": 20, "purchase_price": 18.5, "mrp": 26.0, "selling_price": 24.0, "gst_percent": 5, "rack": "A1", "updated_at": now_iso()},
            {"id": 2, "code": "MED-002", "name": "Azithromycin 500", "category": "Antibiotic", "batch_no": "AZI988", "expiry_date": (today + timedelta(days=240)).isoformat(), "supplier": "Cipla", "unit": "Strip", "quantity": 42, "min_stock": 12, "purchase_price": 82.0, "mrp": 114.0, "selling_price": 108.0, "gst_percent": 12, "rack": "A2", "updated_at": now_iso()},
            {"id": 3, "code": "MED-003", "name": "Pantoprazole", "category": "Tablet", "batch_no": "PAN111", "expiry_date": (today + timedelta(days=180)).isoformat(), "supplier": "Alkem", "unit": "Strip", "quantity": 16, "min_stock": 15, "purchase_price": 54.0, "mrp": 76.0, "selling_price": 71.0, "gst_percent": 12, "rack": "B1", "updated_at": now_iso()},
            {"id": 4, "code": "MED-004", "name": "ORS Sachet", "category": "General", "batch_no": "ORS501", "expiry_date": (today + timedelta(days=540)).isoformat(), "supplier": "FDC", "unit": "Piece", "quantity": 9, "min_stock": 20, "purchase_price": 12.0, "mrp": 20.0, "selling_price": 18.0, "gst_percent": 5, "rack": "C4", "updated_at": now_iso()},
        ],
        "customers": [
            {"id": 1, "name": "Walk-in Customer", "phone": "", "address": "Gopamau Hardoi", "created_at": now_iso()}
        ],
        "sales": [],
        "sale_items": [],
    }


def save_store(store: dict) -> None:
    DATA_PATH.write_text(json.dumps(store, indent=2), encoding="utf-8")


def init_store() -> None:
    if not DATA_PATH.exists():
        save_store(seed_data())
    else:
        store = json.loads(DATA_PATH.read_text(encoding="utf-8"))
        store = normalize_store(store)
        save_store(store)


def normalize_store(store: dict) -> dict:
    defaults = seed_data()
    store.setdefault("next_ids", defaults["next_ids"].copy())
    store.setdefault("pharmacy", defaults["pharmacy"].copy())
    store.setdefault("users", defaults["users"].copy())
    store.setdefault("medicines", [])
    store.setdefault("customers", defaults["customers"].copy())
    store.setdefault("sales", [])
    store.setdefault("sale_items", [])

    if not store["users"]:
        store["users"] = defaults["users"].copy()

    max_med_id = 0
    for med in store["medicines"]:
        med.setdefault("code", f"MED-{int(med.get('id', 0) or 0):03d}" if med.get("id") else "")
        med.setdefault("batch_no", "")
        med.setdefault("mrp", float(med.get("selling_price", 0)))
        med.setdefault("rack", "")
        med.setdefault("supplier", "")
        med.setdefault("unit", "Strip")
        med.setdefault("updated_at", now_iso())
        max_med_id = max(max_med_id, int(med.get("id", 0)))

    max_user_id = 0
    for user in store["users"]:
        user.setdefault("role", "staff")
        user.setdefault("display_name", user.get("username", "User"))
        user.setdefault("phone", "")
        user.setdefault("active", True)
        user.setdefault("created_at", now_iso())
        max_user_id = max(max_user_id, int(user.get("id", 0)))

    store["next_ids"]["medicine"] = max(store["next_ids"].get("medicine", 1), max_med_id + 1)
    store["next_ids"]["user"] = max(store["next_ids"].get("user", 1), max_user_id + 1)
    return store


def load_store() -> dict:
    init_store()
    return normalize_store(json.loads(DATA_PATH.read_text(encoding="utf-8")))


def next_id(store: dict, key: str) -> int:
    current = int(store["next_ids"].get(key, 1))
    store["next_ids"][key] = current + 1
    return current


def current_user() -> dict | None:
    return session.get("user")


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        user = current_user()
        if not user:
            return redirect(url_for("login"))
        if user.get("role") != "admin":
            return jsonify({"ok": False, "message": "Admin access required."}), 403
        return view(*args, **kwargs)
    return wrapped_view


def fetch_dashboard_stats(store: dict) -> dict:
    today_str = date.today().isoformat()
    month_token = today_str[:7]
    low_stock = sum(1 for med in store["medicines"] if int(med["quantity"]) <= int(med["min_stock"]))
    expiring = sum(1 for med in store["medicines"] if med.get("expiry_date") and med["expiry_date"] <= (date.today() + timedelta(days=90)).isoformat())
    today_sales = sum(float(sale["net_total"]) for sale in store["sales"] if sale["created_at"].startswith(today_str))
    monthly_sales = sum(float(sale["net_total"]) for sale in store["sales"] if sale["created_at"].startswith(month_token))
    stock_value = sum(float(med["purchase_price"]) * int(med["quantity"]) for med in store["medicines"])
    sales_by_day = {}
    for sale in store["sales"]:
        label = sale["created_at"][:10]
        sales_by_day[label] = sales_by_day.get(label, 0) + float(sale["net_total"])
    recent_days = sorted(sales_by_day.items())[-7:]
    return {
        "totalMedicines": len(store["medicines"]),
        "lowStockCount": low_stock,
        "expiringCount": expiring,
        "todaySales": round(today_sales, 2),
        "monthlySales": round(monthly_sales, 2),
        "stockValue": round(stock_value, 2),
        "salesChart": [{"label": d[0][8:10] + "/" + d[0][5:7], "amount": round(d[1], 2)} for d in recent_days],
    }


def next_bill_no(store: dict) -> str:
    today_token = datetime.now().strftime("%Y%m%d")
    todays = sum(1 for sale in store["sales"] if sale["created_at"].startswith(date.today().isoformat()))
    return f"MMS-{today_token}-{todays + 1:03d}"


def generate_medicine_code(store: dict) -> str:
    return f"MED-{int(store['next_ids'].get('medicine', 1)):03d}"


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        store = load_store()
        user = next((u for u in store["users"] if u["username"] == username and u["password"] == password and u.get("active", True)), None)
        if user:
            session["user"] = {"id": user["id"], "username": user["username"], "name": user.get("display_name", user["username"]), "role": user.get("role", "staff")}
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


@app.route("/billing")
@login_required
def billing():
    return render_template("billing.html")


@app.route("/inventory")
@login_required
def inventory():
    return render_template("inventory.html")


@app.route("/customers")
@login_required
def customers():
    return render_template("customers.html")


@app.route("/reports")
@login_required
def reports():
    return render_template("reports.html")


@app.route("/settings")
@login_required
def settings():
    return render_template("settings.html")


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    return jsonify(fetch_dashboard_stats(load_store()))


@app.route("/api/medicines")
@login_required
def api_medicines():
    store = load_store()
    search = request.args.get("q", "").strip().lower()
    items = store["medicines"]
    if search:
        items = [m for m in items if search in m.get("name", "").lower() or search in m.get("code", "").lower() or search in m.get("category", "").lower()]
    items = sorted(items, key=lambda item: item.get("updated_at", ""), reverse=True)
    return jsonify(items)


@app.route("/api/medicines", methods=["POST"])
@login_required
def api_add_medicine():
    store = load_store()
    payload = request.get_json(force=True)
    medicine_id = next_id(store, "medicine")
    medicine = {
        "id": medicine_id,
        "code": generate_medicine_code(store),
        "name": payload["name"].strip(),
        "category": payload.get("category", "").strip(),
        "batch_no": "",
        "expiry_date": payload.get("expiry_date", "").strip(),
        "supplier": payload.get("supplier", "").strip(),
        "unit": payload.get("unit", "Strip").strip(),
        "quantity": int(payload.get("quantity", 0)),
        "min_stock": int(payload.get("min_stock", 10)),
        "purchase_price": float(payload.get("purchase_price", 0)),
        "mrp": float(payload.get("mrp", payload.get("selling_price", 0))),
        "selling_price": float(payload.get("selling_price", 0)),
        "gst_percent": float(payload.get("gst_percent", 0)),
        "rack": payload.get("rack", "").strip(),
        "updated_at": now_iso(),
    }
    store["medicines"].append(medicine)
    save_store(store)
    return jsonify({"ok": True})


@app.route("/api/medicines/<int:medicine_id>", methods=["PUT"])
@login_required
def api_update_medicine(medicine_id: int):
    store = load_store()
    payload = request.get_json(force=True)
    for index, item in enumerate(store["medicines"]):
        if int(item["id"]) == medicine_id:
            updated = {
                **item,
                "name": payload["name"].strip(),
                "category": payload.get("category", "").strip(),
                "expiry_date": payload.get("expiry_date", "").strip(),
                "supplier": payload.get("supplier", "").strip(),
                "unit": payload.get("unit", "Strip").strip(),
                "quantity": int(payload.get("quantity", 0)),
                "min_stock": int(payload.get("min_stock", 10)),
                "purchase_price": float(payload.get("purchase_price", 0)),
                "mrp": float(payload.get("mrp", payload.get("selling_price", 0))),
                "selling_price": float(payload.get("selling_price", 0)),
                "gst_percent": float(payload.get("gst_percent", 0)),
                "rack": payload.get("rack", "").strip(),
                "updated_at": now_iso(),
            }
            store["medicines"][index] = updated
            save_store(store)
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "Medicine not found."}), 404


@app.route("/api/medicines/<int:medicine_id>", methods=["DELETE"])
@login_required
def api_delete_medicine(medicine_id: int):
    store = load_store()
    store["medicines"] = [item for item in store["medicines"] if int(item["id"]) != medicine_id]
    save_store(store)
    return jsonify({"ok": True})


@app.route("/api/customers")
@login_required
def api_customers():
    store = load_store()
    q = request.args.get("q", "").strip().lower()
    items = store["customers"]
    if q:
        items = [c for c in items if q in c["name"].lower() or q in c["phone"].lower()]
    items = sorted(items, key=lambda item: item.get("created_at", ""), reverse=True)
    return jsonify(items)


@app.route("/api/customers", methods=["POST"])
@login_required
def api_add_customer():
    store = load_store()
    payload = request.get_json(force=True)
    customer = {"id": next_id(store, "customer"), "name": payload["name"].strip(), "phone": payload.get("phone", "").strip(), "address": payload.get("address", "").strip(), "created_at": now_iso()}
    store["customers"].append(customer)
    save_store(store)
    return jsonify({"ok": True, "id": customer["id"]})


@app.route("/api/reports")
@login_required
def api_reports():
    store = load_store()
    recent_sales = sorted(store["sales"], key=lambda sale: sale["created_at"], reverse=True)[:25]
    movement = {}
    for item in store["sale_items"]:
        row = movement.setdefault(item["medicine_name"], {"medicine_name": item["medicine_name"], "sold_qty": 0, "revenue": 0.0})
        row["sold_qty"] += int(item["quantity"])
        row["revenue"] += float(item["line_total"])
    top_items = sorted(movement.values(), key=lambda row: (row["sold_qty"], row["revenue"]), reverse=True)[:10]
    for item in top_items:
        item["revenue"] = round(item["revenue"], 2)
    return jsonify({"recentSales": recent_sales, "topItems": top_items})


@app.route("/api/sales", methods=["POST"])
@login_required
def api_create_sale():
    store = load_store()
    payload = request.get_json(force=True)
    items = payload.get("items", [])
    if not items:
        return jsonify({"ok": False, "message": "Add at least one medicine to create a bill."}), 400

    inventory = {int(item["id"]): item for item in store["medicines"]}
    for item in items:
        med = inventory.get(int(item["medicine_id"]))
        if med is None:
            return jsonify({"ok": False, "message": "Medicine not found."}), 404
        if int(med["quantity"]) < int(item["quantity"]):
            return jsonify({"ok": False, "message": f"Insufficient stock for {med['name']}. Available: {med['quantity']}"}), 400

    bill_no = next_bill_no(store)
    sale = {
        "id": next_id(store, "sale"),
        "bill_no": bill_no,
        "customer_id": payload.get("customer_id"),
        "customer_name": payload.get("customer_name", "Walk-in Customer").strip() or "Walk-in Customer",
        "gross_total": float(payload.get("gross_total", 0)),
        "gst_total": float(payload.get("gst_total", 0)),
        "discount_total": float(payload.get("discount_total", 0)),
        "net_total": float(payload.get("net_total", 0)),
        "payment_mode": payload.get("payment_mode", "Cash"),
        "created_at": now_iso(),
    }
    store["sales"].append(sale)

    for item in items:
        med = inventory[int(item["medicine_id"])]
        med["quantity"] = int(med["quantity"]) - int(item["quantity"])
        med["updated_at"] = now_iso()
        store["sale_items"].append({
            "id": next_id(store, "sale_item"),
            "sale_id": sale["id"],
            "medicine_id": int(item["medicine_id"]),
            "medicine_name": item["name"],
            "quantity": int(item["quantity"]),
            "unit_price": float(item["price"]),
            "gst_percent": float(item["gst_percent"]),
            "discount_percent": float(item.get("discount_percent", 0)),
            "line_total": float(item["line_total"]),
        })

    save_store(store)
    return jsonify({"ok": True, "bill_no": bill_no})


@app.route("/api/profile")
@login_required
def api_profile():
    store = load_store()
    user = current_user()
    full_user = next((u for u in store["users"] if int(u["id"]) == int(user["id"])), None)
    return jsonify({"user": full_user, "pharmacy": store["pharmacy"]})


@app.route("/api/profile", methods=["PUT"])
@login_required
def api_update_profile():
    store = load_store()
    payload = request.get_json(force=True)
    user = current_user()
    full_user = next((u for u in store["users"] if int(u["id"]) == int(user["id"])), None)
    if not full_user:
        return jsonify({"ok": False, "message": "User not found."}), 404
    new_username = payload.get("username", full_user["username"]).strip()
    if any(u["username"] == new_username and int(u["id"]) != int(full_user["id"]) for u in store["users"]):
        return jsonify({"ok": False, "message": "Username already exists."}), 400
    full_user["username"] = new_username
    full_user["display_name"] = payload.get("display_name", full_user["display_name"]).strip()
    full_user["phone"] = payload.get("phone", full_user.get("phone", "")).strip()
    store["pharmacy"]["name"] = payload.get("pharmacy_name", store["pharmacy"]["name"]).strip()
    store["pharmacy"]["address"] = payload.get("pharmacy_address", store["pharmacy"]["address"]).strip()
    store["pharmacy"]["phone"] = payload.get("pharmacy_phone", store["pharmacy"].get("phone", "")).strip()
    store["pharmacy"]["owner_name"] = payload.get("owner_name", store["pharmacy"].get("owner_name", "")).strip()
    save_store(store)
    session["user"]["name"] = full_user["display_name"]
    session["user"]["username"] = full_user["username"]
    return jsonify({"ok": True})


@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    store = load_store()
    payload = request.get_json(force=True)
    user = current_user()
    full_user = next((u for u in store["users"] if int(u["id"]) == int(user["id"])), None)
    if not full_user:
        return jsonify({"ok": False, "message": "User not found."}), 404
    if full_user["password"] != payload.get("current_password", ""):
        return jsonify({"ok": False, "message": "Current password is incorrect."}), 400
    new_password = payload.get("new_password", "").strip()
    if len(new_password) < 4:
        return jsonify({"ok": False, "message": "New password must be at least 4 characters."}), 400
    full_user["password"] = new_password
    save_store(store)
    return jsonify({"ok": True})


@app.route("/api/users")
@login_required
def api_users():
    store = load_store()
    current = current_user()
    if current.get("role") != "admin":
        return jsonify([])
    users = [{k: v for k, v in user.items() if k != "password"} for user in store["users"]]
    return jsonify(users)


@app.route("/api/users", methods=["POST"])
@admin_required
def api_add_user():
    store = load_store()
    payload = request.get_json(force=True)
    username = payload.get("username", "").strip()
    if any(u["username"] == username for u in store["users"]):
        return jsonify({"ok": False, "message": "Username already exists."}), 400
    user = {
        "id": next_id(store, "user"),
        "username": username,
        "password": payload.get("password", "1234").strip(),
        "role": payload.get("role", "staff").strip(),
        "display_name": payload.get("display_name", username).strip(),
        "phone": payload.get("phone", "").strip(),
        "active": True,
        "created_at": now_iso(),
    }
    store["users"].append(user)
    save_store(store)
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def api_update_user(user_id: int):
    store = load_store()
    payload = request.get_json(force=True)
    for user in store["users"]:
        if int(user["id"]) == user_id:
            new_username = payload.get("username", user["username"]).strip()
            if any(u["username"] == new_username and int(u["id"]) != int(user_id) for u in store["users"]):
                return jsonify({"ok": False, "message": "Username already exists."}), 400
            user["username"] = new_username
            user["display_name"] = payload.get("display_name", user["display_name"]).strip()
            user["phone"] = payload.get("phone", user.get("phone", "")).strip()
            user["role"] = payload.get("role", user.get("role", "staff")).strip()
            user["active"] = bool(payload.get("active", True))
            if payload.get("password"):
                user["password"] = payload["password"].strip()
            save_store(store)
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "User not found."}), 404


@app.route("/api/export/excel")
@login_required
def api_export_excel():
    if Workbook is None:
        return jsonify({"ok": False, "message": "Install openpyxl to enable Excel export."}), 500

    store = load_store()
    workbook = Workbook()
    ws_meds = workbook.active
    ws_meds.title = "Inventory"
    ws_meds.append(["Name", "Category", "Expiry Date", "Supplier", "Unit", "Quantity", "Min Stock", "Purchase Price", "MRP", "Selling Price", "GST %", "Rack"])
    for med in store["medicines"]:
        ws_meds.append([med["name"], med["category"], med["expiry_date"], med["supplier"], med["unit"], med["quantity"], med["min_stock"], med["purchase_price"], med.get("mrp", med["selling_price"]), med["selling_price"], med["gst_percent"], med["rack"]])

    ws_sales = workbook.create_sheet("Sales")
    ws_sales.append(["Bill No", "Customer", "Gross Total", "GST Total", "Discount Total", "Net Total", "Payment Mode", "Created At"])
    for sale in store["sales"]:
        ws_sales.append([sale["bill_no"], sale["customer_name"], sale["gross_total"], sale["gst_total"], sale["discount_total"], sale["net_total"], sale["payment_mode"], sale["created_at"]])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"mirza_medical_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/import/excel", methods=["POST"])
@login_required
def api_import_excel():
    if load_workbook is None:
        return jsonify({"ok": False, "message": "Install openpyxl to enable Excel import."}), 500

    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"ok": False, "message": "Choose an Excel file first."}), 400

    workbook = load_workbook(uploaded, data_only=True)
    if "Inventory" not in workbook.sheetnames:
        return jsonify({"ok": False, "message": "Excel file must contain an Inventory sheet."}), 400

    rows = list(workbook["Inventory"].iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"ok": False, "message": "Inventory sheet is empty."}), 400

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    store = load_store()
    store["medicines"] = []
    store["next_ids"]["medicine"] = 1
    for row_values in rows[1:]:
        if not any(row_values):
            continue
        record = dict(zip(headers, row_values))
        medicine_id = next_id(store, "medicine")
        store["medicines"].append({
            "id": medicine_id,
            "code": f"MED-{medicine_id:03d}",
            "name": str(record.get("Name") or record.get("Medicine Name") or "").strip(),
            "category": str(record.get("Category") or "").strip(),
            "batch_no": "",
            "expiry_date": str(record.get("Expiry Date") or "").strip(),
            "supplier": str(record.get("Supplier") or "").strip(),
            "unit": str(record.get("Unit") or "Strip").strip(),
            "quantity": int(record.get("Quantity") or 0),
            "min_stock": int(record.get("Min Stock") or 10),
            "purchase_price": float(record.get("Purchase Price") or 0),
            "mrp": float(record.get("MRP") or record.get("Mrp") or record.get("Selling Price") or 0),
            "selling_price": float(record.get("Selling Price") or 0),
            "gst_percent": float(record.get("GST %") or record.get("GST") or 0),
            "rack": str(record.get("Rack") or "").strip(),
            "updated_at": now_iso(),
        })
    save_store(store)
    return jsonify({"ok": True, "message": f"Imported {len(store['medicines'])} medicines successfully."})


if __name__ == "__main__":
    init_store()
    app.run(debug=True)
